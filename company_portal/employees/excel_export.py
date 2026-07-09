"""
Module 4: Excel Export

Builds openpyxl workbooks for:
 * Employee List
 * Department List
 * Salary Report
 * Attendance Report

Each function returns an io.BytesIO buffer ready to stream in an HttpResponse.
"""
import io
from datetime import datetime, date

from django.db.models import Avg, Count
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')


def _autosize(ws):
    for i, column_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(max_len + 3, 45))


def _to_buffer(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_employee_list(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        "Employee ID", "First Name", "Last Name", "Email", "Phone",
        "Department", "Designation", "Salary", "Joining Date", "Status",
    ]
    ws.append(headers)
    for emp in queryset:
        ws.append([
            emp.employee_id, emp.first_name, emp.last_name, emp.email, emp.phone,
            emp.department.name if emp.department else '',
            emp.designation, float(emp.salary),
            emp.joining_date.strftime('%Y-%m-%d') if emp.joining_date else '',
            emp.status,
        ])
    _style_header(ws)
    _autosize(ws)
    return _to_buffer(wb)


def export_department_list(department_queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Departments"
    ws.append(["Department", "Employee Count"])
    qs = department_queryset.annotate(employee_count=Count('employees'))
    for dept in qs:
        ws.append([dept.name, dept.employee_count])
    _style_header(ws)
    _autosize(ws)
    return _to_buffer(wb)


def export_salary_report(employee_queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Report"
    ws.append(["Employee ID", "Name", "Department", "Designation", "Salary"])
    total = 0
    for emp in employee_queryset:
        ws.append([
            emp.employee_id, emp.full_name,
            emp.department.name if emp.department else '',
            emp.designation, float(emp.salary),
        ])
        total += float(emp.salary)
    ws.append([])
    ws.append(["", "", "", "Total Payroll", total])
    _style_header(ws)
    _autosize(ws)

    # Department summary sheet
    ws2 = wb.create_sheet("Department Summary")
    ws2.append(["Department", "Employee Count", "Average Salary", "Total Salary"])
    from .models import Department
    for dept in Department.objects.all():
        emps = employee_queryset.filter(department=dept)
        count = emps.count()
        avg = emps.aggregate(avg=Avg('salary'))['avg'] or 0
        dept_total = sum(float(e.salary) for e in emps)
        ws2.append([dept.name, count, round(float(avg), 2), dept_total])
    _style_header(ws2)
    _autosize(ws2)

    return _to_buffer(wb)


EXPECTED_IMPORT_HEADERS = [
    "Employee ID", "First Name", "Last Name", "Email", "Phone",
    "Department", "Designation", "Salary", "Joining Date",
]


def _coerce_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    value = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognised date format: {value}")


def import_employees_excel(uploaded_file):
    """
    Module 3: Excel Import
    Reads the uploaded .xlsx, validates every row, skips invalid/duplicate
    records, saves valid employees, and returns an import summary dict:
    {total_rows, created, failed, errors: [{row, error}]}
    """
    from .models import Department, Employee  # local import avoids circularity

    try:
        wb = load_workbook(filename=io.BytesIO(uploaded_file.read()), data_only=True)
    except Exception as exc:
        return {"total_rows": 0, "created": 0, "failed": 0,
                "errors": [{"row": 0, "error": f"Could not read Excel file: {exc}"}]}

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {"total_rows": 0, "created": 0, "failed": 0,
                "errors": [{"row": 0, "error": "Sheet is empty."}]}

    header_row = [str(h).strip() if h is not None else '' for h in header_row]
    missing = [h for h in EXPECTED_IMPORT_HEADERS if h not in header_row]
    if missing:
        return {"total_rows": 0, "created": 0, "failed": 0,
                "errors": [{"row": 0, "error": f"Missing required columns: {', '.join(missing)}"}]}

    col_index = {name: header_row.index(name) for name in EXPECTED_IMPORT_HEADERS}

    seen_ids, seen_emails = set(), set()
    total_rows, created, failed = 0, 0, 0
    errors = []

    for excel_row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell in (None, '') for cell in row):
            continue  # skip fully blank rows, don't count them
        total_rows += 1
        try:
            def get(col_name):
                idx = col_index[col_name]
                return row[idx] if idx < len(row) else None

            emp_id = str(get("Employee ID") or '').strip()
            email = str(get("Email") or '').strip().lower()

            if not emp_id or not email:
                raise ValueError("Employee ID and Email are required.")
            if emp_id in seen_ids or email in seen_emails:
                raise ValueError("Duplicate row within uploaded file.")
            if Employee.objects.filter(employee_id=emp_id).exists():
                raise ValueError(f"Employee ID '{emp_id}' already exists.")
            if Employee.objects.filter(email__iexact=email).exists():
                raise ValueError(f"Email '{email}' already exists.")

            salary_raw = get("Salary")
            try:
                salary = float(salary_raw) if salary_raw not in (None, '') else 0
            except (TypeError, ValueError):
                raise ValueError(f"Invalid salary value: {salary_raw}")

            department = None
            dept_name = str(get("Department") or '').strip()
            if dept_name:
                department, _ = Department.objects.get_or_create(name=dept_name)

            Employee.objects.create(
                employee_id=emp_id,
                first_name=str(get("First Name") or '').strip(),
                last_name=str(get("Last Name") or '').strip(),
                email=email,
                phone=str(get("Phone") or '').strip(),
                department=department,
                designation=str(get("Designation") or '').strip(),
                salary=salary,
                joining_date=_coerce_date(get("Joining Date")),
            )
            seen_ids.add(emp_id)
            seen_emails.add(email)
            created += 1
        except Exception as exc:
            failed += 1
            errors.append({"row": excel_row_num, "error": str(exc)})

    return {"total_rows": total_rows, "created": created, "failed": failed, "errors": errors}


def export_attendance_report(attendance_queryset):

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    ws.append(["Employee ID", "Name", "Date", "Status"])
    for att in attendance_queryset.select_related('employee'):
        ws.append([att.employee.employee_id, att.employee.full_name,
                   att.date.strftime('%Y-%m-%d'), att.status])
    _style_header(ws)
    _autosize(ws)
    return _to_buffer(wb)
