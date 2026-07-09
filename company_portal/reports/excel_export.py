"""
Module 10: Reports Module - Excel (.xlsx) report builders.
"""
import io

from django.db.models import Avg, Count
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from employees.models import Attendance, Department, Employee

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')


def _autosize(ws):
    for i, column_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(max_len + 3, 45))


def _to_buffer(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_employee_report_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Active Employees"
    ws.append(["Employee ID", "Name", "Department", "Designation", "Joining Date"])
    for emp in Employee.objects.filter(status='ACTIVE').select_related('department'):
        ws.append([emp.employee_id, emp.full_name,
                   emp.department.name if emp.department else '', emp.designation,
                   emp.joining_date.strftime('%Y-%m-%d') if emp.joining_date else ''])
    _style_header(ws)
    _autosize(ws)

    ws2 = wb.create_sheet("Inactive Employees")
    ws2.append(["Employee ID", "Name", "Department", "Designation", "Joining Date"])
    for emp in Employee.objects.filter(status='INACTIVE').select_related('department'):
        ws2.append([emp.employee_id, emp.full_name,
                    emp.department.name if emp.department else '', emp.designation,
                    emp.joining_date.strftime('%Y-%m-%d') if emp.joining_date else ''])
    _style_header(ws2)
    _autosize(ws2)
    return _to_buffer(wb)


def build_department_report_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Department Report"
    ws.append(["Department", "Employee Count", "Average Salary"])
    for dept in Department.objects.annotate(employee_count=Count('employees'), avg_salary=Avg('employees__salary')):
        ws.append([dept.name, dept.employee_count, round(float(dept.avg_salary or 0), 2)])
    _style_header(ws)
    _autosize(ws)
    return _to_buffer(wb)


def build_salary_report_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Payroll"
    ws.append(["Employee ID", "Name", "Department", "Salary"])
    total = 0
    for emp in Employee.objects.select_related('department').all():
        ws.append([emp.employee_id, emp.full_name,
                   emp.department.name if emp.department else '', float(emp.salary)])
        total += float(emp.salary)
    ws.append(["", "", "Total Monthly Payroll", total])
    _style_header(ws)
    _autosize(ws)

    ws2 = wb.create_sheet("Department Salary Summary")
    ws2.append(["Department", "Total Salary", "Average Salary"])
    for dept in Department.objects.annotate(total_salary=Count('employees')):
        emps = Employee.objects.filter(department=dept)
        total_salary = sum(float(e.salary) for e in emps)
        avg_salary = total_salary / emps.count() if emps.count() else 0
        ws2.append([dept.name, total_salary, round(avg_salary, 2)])
    _style_header(ws2)
    _autosize(ws2)
    return _to_buffer(wb)


def build_attendance_report_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    ws.append(["Employee ID", "Name", "Present", "Absent", "Leave"])
    for emp in Employee.objects.all():
        present = Attendance.objects.filter(employee=emp, status='PRESENT').count()
        absent = Attendance.objects.filter(employee=emp, status='ABSENT').count()
        leave = Attendance.objects.filter(employee=emp, status='LEAVE').count()
        ws.append([emp.employee_id, emp.full_name, present, absent, leave])
    _style_header(ws)
    _autosize(ws)
    return _to_buffer(wb)


def build_dashboard_excel(dashboard_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Metric", "Value"])
    ws.append(["Total Employees", dashboard_data["employee_count"]])
    ws.append(["Active Employees", dashboard_data["active_employees"]])
    ws.append(["Inactive Employees", dashboard_data["inactive_employees"]])
    ws.append(["Total Departments", dashboard_data["department_count"]])
    ws.append(["Total Monthly Payroll", dashboard_data["salary_statistics"]["total_payroll"]])
    ws.append(["Average Salary", dashboard_data["salary_statistics"]["average_salary"]])
    ws.append(["Present Today", dashboard_data["attendance_summary"]["present"]])
    ws.append(["Absent Today", dashboard_data["attendance_summary"]["absent"]])
    ws.append(["On Leave Today", dashboard_data["attendance_summary"]["leave"]])
    _style_header(ws)
    _autosize(ws)
    return _to_buffer(wb)
